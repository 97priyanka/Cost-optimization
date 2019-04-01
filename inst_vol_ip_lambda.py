import boto3
import openpyxl
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart

def lambda_handler(event, context):
    ses = boto3.client('ses')
    wb = openpyxl.Workbook()
    inst = wb.active
    inst.title="ec2_stopped_inst"
    content=['Name','Instance Id','Ip Address','Instance Type','Availability Zone','State Transition Reason','State Reason Message']
    r=1
    c=1
    for item in content:
        cel = inst.cell(row = r, column = c)
        cel.value=item
        c+=1
    ec2 = boto3.client('ec2')
    region=ec2.describe_regions()
    for reg in region['Regions']:
        ec2=boto3.client('ec2',region_name=reg['RegionName'])
        response = ec2.describe_instances()
        for reservation in response["Reservations"]:
            for instance in reservation["Instances"]:
                for tag in instance['Tags']:
                    if(tag['Key']=='Name'):
                        instance_name=tag['Value']
                state=instance['State']
                if(state['Name']=='stopped'):
                    reason=instance['StateReason']
                    az=instance['Placement']
                    content=[instance_name,instance['InstanceId'],instance['PrivateIpAddress'],instance['InstanceType'],az['AvailabilityZone'],instance['StateTransitionReason'],reason['Message']]
                    r+=1
                    c=1
                    for item in content:
                        cel = inst.cell(row = r, column = c)
                        cel.value=item
                        c+=1

    volu=wb.create_sheet(index = 1 , title = "ebs_available_volume")
    content=['Name','Volume Id','Size','Volume Type','IOPS','Created on','Availability Zone']
    r=1
    c=1
    for item in content:
        cel = volu.cell(row = r, column = c)
        cel.value=item
        c+=1
    region=ec2.describe_regions()
    for reg in region['Regions']:
        ec2=boto3.client('ec2',region_name=reg['RegionName'])
        volumes=ec2.describe_volumes()
        for vol in volumes['Volumes']:
            tags=ec2.describe_tags(Filters=[{'Name': 'resource-id','Values': [vol['VolumeId']]}])
            for tag in tags['Tags']:
                if tag['Key']=='Name':
                    name=tag['Value']
            date=str(vol['CreateTime'])
            if vol['State']=='available':
                content=[name,vol['VolumeId'],vol['Size'],vol['VolumeType'],vol['Iops'],date,vol['AvailabilityZone']]
                r+=1
                c=1
                for item in content:
                    cel = volu.cell(row = r, column = c)
                    cel.value=item
                    c+=1

    ip = wb.create_sheet(index = 2 , title = "unused_elastic_ip")
    r=1
    cel=ip.cell(row = r , column = 1)
    cel.value='Elastic Ip'
    region=ec2.describe_regions()
    for reg in region['Regions']:
        ec2=boto3.client('ec2',region_name=reg['RegionName'])
        elastic_ips=ec2.describe_addresses()
        for ip in elastic_ips['Addresses']:
            info_ip=str(ip)
            result=info_ip.find('NetworkInterfaceId')
            if result<0:
                r+=1
                cel = ip.cell(row = r , column = 1)
                cel.value=ip['PublicIp']
    wb.save('/tmp/excell.xlsx')

    msg=MIMEMultipart()
    msg['Subject']='Information of stopped ec2 instances and ebs volume'
    msg['From']='-----------senders-mail-id------------'
    msg['To']='--------------receivers-mail-id------------'
    part = MIMEText('This is the data of stopped EC2 Instances, available EBS Volumes and unassociated Elastic IP')
    msg.attach(part)
    part= MIMEApplication(open('/tmp/excell.xlsx','rb').read())
    part.add_header('Content-Disposition', 'attachment', filename='Scripts.xlsx')
    msg.attach(part)
    ses.send_raw_email(RawMessage={'Data':msg.as_string()},Source=msg['From'],Destinations=[msg['To']])

